#!/usr/bin/env python3
"""Lê o Excel consolidado e regenera toda a base de dados do CapacitIA Servidores.

Pipeline:
  1. Extrai 2025 DADOS + 2026 DADOS do Excel
  2. Normaliza nomes de colunas (internal format)
  3. Unifica colunas OUTROS (cargo + cargo_outros, orgao + orgao_outros, vinculo + vinculo_outros)
  4. Aplica normalização canônica (órgão, cargo, vínculo)
  5. Infere órgão externo
  6. Salva CSV em .data/raw/ (com colunas no formato original para compatibilidade)
  7. Reexecuta CapacitiaCSVProcessor.process_all() para regenerar todos os parquets
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import numpy as np
import logging
import openpyxl

from src.process_csv_to_parquet import CapacitiaCSVProcessor
import src.config as cfg

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_PATH = PROJECT_ROOT / "archive" / "relatorio" / "capacitia-dados.xlsx"
RAW_PATH = PROJECT_ROOT / ".data" / "raw"
PROCESSED_PATH = PROJECT_ROOT / ".data" / "processed"

# Ordem e nomes das colunas no CSV final (compatível com o pipeline existente)
CSV_COLUMNS = [
    "ANO", "EVENTO", "FORMATO", "ÓRGÃO EXTERNO", "EIXO",
    "LOCAL DE REALIZAÇÃO", "NOME",
    "CARGO", "CARGO OUTROS",
    "ÓRGÃO", "ÓRGÃO OUTROS",
    "VÍNCULO", "VÍNCULO OUTROS",
    "CERTIFICADO", "CARGO DE GESTÃO", "SERVIDOR DO ESTADO",
]

# Mapeamento: nome interno → nome CSV
INTERNAL_TO_CSV = {
    "ano": "ANO",
    "evento": "EVENTO",
    "formato": "FORMATO",
    "orgao_externo": "ÓRGÃO EXTERNO",
    "eixo": "EIXO",
    "local_realizacao": "LOCAL DE REALIZAÇÃO",
    "nome": "NOME",
    "cargo": "CARGO",
    "cargo_outros": "CARGO OUTROS",
    "orgao": "ÓRGÃO",
    "orgao_outros": "ÓRGÃO OUTROS",
    "vinculo": "VÍNCULO",
    "vinculo_outros": "VÍNCULO OUTROS",
    "certificado": "CERTIFICADO",
    "cargo_gestao": "CARGO DE GESTÃO",
    "servidor_estado": "SERVIDOR DO ESTADO",
}


def _normalize_col(name: str) -> str:
    return (
        name.strip()
        .lower()
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("õ", "o")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace(" ", "_")
        .replace("___", "_")
        .replace("__", "_")
    )


def _find_header_row(ws) -> tuple[int, list[str]]:
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        if "evento" in cells[0].lower() and "formato" in " ".join(cells[:5]).lower():
            return i, cells
    raise ValueError("Cabeçalho não encontrado")


def _read_sheet(ws, year: str) -> pd.DataFrame:
    header_row_idx, raw_headers = _find_header_row(ws)
    headers_norm = [_normalize_col(h) for h in raw_headers]

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells or not cells[0]:
            continue
        if cells[0].lower() in ("", "nan", "none"):
            continue
        rows.append(cells[: len(headers_norm)])

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=headers_norm)

    for col in df.columns:
        df[col] = df[col].replace(["nan", "None", ""], pd.NA)

    df = df.dropna(how="all").reset_index(drop=True)
    df["ano"] = year

    # Normalizar nome da coluna cargo_outros (2026 usa "cargo_outro" sem s)
    if "cargo_outro" in df.columns and "cargo_outros" not in df.columns:
        df.rename(columns={"cargo_outro": "cargo_outros"}, inplace=True)

    return df


def _ensure_outros_cols(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["cargo_outros", "orgao_outros", "vinculo_outros"]:
        if col not in df.columns:
            df[col] = ""
    return df


def process_excel():
    logger.info("Lendo Excel: %s", EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH)

    sheet_year_map = {
        "2025 DADOS": "2025",
        "2026 DADOS": "2026",
    }

    all_dfs = []
    for sheet_name, year in sheet_year_map.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("Planilha %s não encontrada, pulando", sheet_name)
            continue

        logger.info("Lendo %s (ano %s)...", sheet_name, year)
        df = _read_sheet(wb[sheet_name], year)
        logger.info("  → %d registros lidos", len(df))
        all_dfs.append(df)

    wb.close()

    if not all_dfs:
        logger.error("Nenhum dado encontrado!")
        return

    df = pd.concat(all_dfs, ignore_index=True)
    logger.info("Total consolidado: %d registros", len(df))

    # Garantir colunas OUTROS existam
    df = _ensure_outros_cols(df)

    # --- Unificar OUTROS com as colunas principais ---
    for primary, outros_name in [
        ("cargo", "cargo_outros"),
        ("orgao", "orgao_outros"),
        ("vinculo", "vinculo_outros"),
    ]:
        if primary not in df.columns:
            continue
        p = df[primary].astype(str).str.strip()
        o = df[outros_name].astype(str).str.strip()
        merged = np.where(
            p.str.lower().isin(["", "nan", "outro", "outros"]),
            o,
            p,
        )
        df[primary] = merged

    # Aplicar normalização canônica
    if "orgao" in df.columns:
        df["orgao"] = df["orgao"].map(cfg.canonical_orgao)
    if "cargo" in df.columns:
        df["cargo"] = df["cargo"].map(cfg.canonical_cargo)
    if "vinculo" in df.columns:
        df["vinculo"] = df["vinculo"].map(cfg.canonical_vinculo)

    # Inferir órgão externo
    df["orgao_externo"] = (
        df["orgao"]
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(cfg.ORGAOS_EXTERNOS)
        .map({True: "Sim", False: "Não"})
    )

    # --- Montar DataFrame no formato CSV original ---
    csv_data = {}
    for internal_name, csv_name in INTERNAL_TO_CSV.items():
        if internal_name in df.columns:
            csv_data[csv_name] = df[internal_name].fillna("")
        else:
            csv_data[csv_name] = [""] * len(df)

    df_csv = pd.DataFrame(csv_data)
    df_csv = df_csv[CSV_COLUMNS]

    # Salvar CSV
    RAW_PATH.mkdir(parents=True, exist_ok=True)
    PROCESSED_PATH.mkdir(parents=True, exist_ok=True)
    csv_path = RAW_PATH / "dados_gerais_capacitia.csv"
    logger.info("Salvando CSV: %s", csv_path)
    df_csv.to_csv(csv_path, sep=";", index=False, encoding="utf-8")
    logger.info("CSV salvo com %d linhas e %d colunas", len(df_csv), len(df_csv.columns))

    # Regenerar parquets (Servidores)
    logger.info("Regenerando parquets via CapacitiaCSVProcessor...")
    processor = CapacitiaCSVProcessor(base_path=PROJECT_ROOT)
    processor.process_all()

    # Processar módulos Saúde e Autonomia Digital
    logger.info("Processando módulos adicionais...")
    from src.processors.processor_saude import process_saude
    from src.processors.processor_autonomiadigital_inscricoes import process_autonomiadigital_inscricoes
    from src.processors.processors_autonomiadigital_avaliacoes import process_autonomiadigital_avaliacoes

    erros = []
    for nome, fn in [
        ("saude", process_saude),
        ("autonomia_inscricoes", process_autonomiadigital_inscricoes),
        ("autonomia_avaliacoes", process_autonomiadigital_avaliacoes),
    ]:
        try:
            fn(RAW_PATH, PROCESSED_PATH)
            logger.info("  %s: OK", nome)
        except FileNotFoundError:
            logger.warning("  %s: arquivo não encontrado (pulando)", nome)
        except Exception as e:
            logger.error("  %s: erro - %s", nome, e)
            erros.append(nome)

    # Validar
    if erros:
        logger.warning("Processamento concluído com erros em: %s", ", ".join(erros))
    else:
        logger.info("Todos os módulos processados com sucesso!")

    df_parquet = pd.read_parquet(PROCESSED_PATH / "dados.parquet")
    logger.info("dados.parquet: %d registros, %d colunas", len(df_parquet), len(df_parquet.columns))
    for ano in ["2025", "2026"]:
        qtd = len(df_parquet[df_parquet["ano"] == ano])
        eventos = df_parquet[df_parquet["ano"] == ano]["evento"].nunique()
        logger.info("  %s: %d registros, %d eventos", ano, qtd, eventos)


if __name__ == "__main__":
    process_excel()
